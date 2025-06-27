import boto3
from datetime import date, datetime, timedelta
import csv
import time
import openpyxl
from botocore.exceptions import ClientError

def main():
    # Assume the IAM Automation Role to start
    sts_client = boto3.client('sts')
    assumed_role_object = sts_client.assume_role(
        RoleArn='arn:aws:iam::528150397796:role/528150397796-Devops-Admin-Role',
        RoleSessionName="iamautomation"
    )
    credentials = assumed_role_object['Credentials']
    client = boto3.client(
        'sts',
        aws_access_key_id=credentials['AccessKeyId'],
        aws_secret_access_key=credentials['SecretAccessKey'],
        aws_session_token=credentials['SessionToken']
    )
    
    # Get the root account number
    root_acc = client.get_caller_identity().get('Account')
    print("Root Account:", root_acc)
    assumerole_iam(client, sts_client)

def assumerole_iam(client, sts_client):
    # Load the Excel file and sheet
    csv_filename = "aws_storage_usage.csv"
    storage_data = []
    fname = 'roles.xlsx'
    wb = openpyxl.load_workbook(fname)
    sheet = wb['Sheet1']
    
    # Loop through the specified range of cells for role ARNs
    for row in sheet['A1':'A110']:
        for cell in row:
            role_arn = cell.value
            if role_arn:
                try:
                    # Assume role for each ARN found in the Excel sheet
                    assumed_role_object = client.assume_role(RoleArn=role_arn,RoleSessionName="AssumeRoleSession1")
                    credentials = assumed_role_object['Credentials']
                    
                    iam_client = boto3.client('iam',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])                    
                    sts_client1 = boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
                    s3_client = boto3.client('s3',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
                    ec2_client = boto3.client('ec2',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'],region_name='us-east-1')
                    rds_client = boto3.client('rds',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'],region_name='us-east-1')
                    dynamodb_client = boto3.client('dynamodb',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'],region_name='us-east-1')
                    efs_client = boto3.client('efs',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'],region_name='us-east-1')
                    cw_client = boto3.client('cloudwatch',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'],region_name='us-east-1')
    				
                    account_num = sts_client1.get_caller_identity().get('Account')
                    print(f"Switching to Account: {account_num}")
                    
                    #account_id = get_account_id(sts_client1)
            
                    # Fetch storage sizes
                    s3_size = get_s3_bucket_sizes(s3_client,cw_client)
                    ebs_size = get_ebs_size(ec2_client)
                    rds_size = get_rds_size(rds_client)
                    dynamodb_size = get_dynamodb_size(dynamodb_client)
                    efs_size = get_efs_size(efs_client)
                    #bucket_sizes, total_s3_gb = get_s3_bucket_sizes(sts_client1)
                    
                    # Compute total storage size
                    total_size = s3_size + ebs_size + rds_size + dynamodb_size + efs_size

                    # Store results
                    storage_data.append([account_num, account_num, int(s3_size), int(ebs_size), int(rds_size), int(dynamodb_size), int(efs_size), int(total_size)])

                    print(f"✅ Data collected for (Account {account_num}): {int(total_size)} GB")

                except Exception as e:
                    print(f"❌ Error processing (Account {account_num}): {e}")

            # Avoid hitting AWS API limits
            time.sleep(1)  # Adding delay to prevent throttling

    # Save data to CSV
    with open(csv_filename, mode="w", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(["AWS Account ID", "AWS Profile", "S3 (GB)", "EBS (GB)", "RDS (GB)", "DynamoDB (GB)", "EFS (GB)", "Total (GB)"])
        writer.writerows(storage_data)

def get_account_id(session):
    try:
        return session.client("sts").get_caller_identity().get('Account')
    except:
        return "Unknown"

# Function to get S3 Storage Size
def get_s3_size(s3_client):
    total_size = 0
    try:
        response = s3_client.list_buckets()
        for bucket in response["Buckets"]:
            bucket_name = bucket["Name"]
            try:
                paginator = s3_client.get_paginator("list_objects_v2")
                for page in paginator.paginate(Bucket=bucket_name):
                    if "Contents" in page:
                        for obj in page["Contents"]:
                            total_size += obj["Size"]
            except Exception as e:
                print(f"Error accessing bucket {bucket_name}: {e}")
                continue
        # Convert bytes to GB
        return total_size / (1024 ** 3)
    except:
        return 0

# Function to get EBS Storage Size
def get_ebs_size(ec2_client):
    try:
        volumes = ec2_client.describe_volumes()["Volumes"]
        return sum(volume["Size"] for volume in volumes)  # GB
    except:
        return 0

# Function to get RDS Storage Size
def get_rds_size(rds_client):
    try:
        instances = rds_client.describe_db_instances()["DBInstances"]
        return sum(instance["AllocatedStorage"] for instance in instances)  # GB
    except:
        return 0

# Function to get DynamoDB Storage Size
def get_dynamodb_size(dynamodb_client):
    try:
        tables = dynamodb_client.list_tables()["TableNames"]
        total_size = 0
        for table_name in tables:
            try:
                table_info = dynamodb_client.describe_table(TableName=table_name)
                total_size += table_info["Table"]["TableSizeBytes"]
            except:
                continue
        return int(total_size / (1024 ** 3))  # Convert Bytes to GB
    except:
        return 0

# Function to get EFS Storage Size
def get_efs_size(efs_client):
    try:
        file_systems = efs_client.describe_file_systems()["FileSystems"]
        return int(sum(fs["SizeInBytes"]["Value"] for fs in file_systems) / (1024 ** 3))  # Convert Bytes to GB
    except:
        return 0

# Function to get S3 Bucket Sizes using CloudWatch
def get_s3_bucket_sizes(s3_client, cw_client):

    total_bytes = 0
    bucket_sizes = {}

    try:
        buckets = s3_client.list_buckets()['Buckets']
        for bucket in buckets:
            bucket_name = bucket['Name']
            try:
                response = cw_client.get_metric_statistics(
                    Namespace='AWS/S3',
                    MetricName='BucketSizeBytes',
                    Dimensions=[
                        {'Name': 'BucketName', 'Value': bucket_name},
                        {'Name': 'StorageType', 'Value': 'StandardStorage'}
                    ],
                    StartTime=datetime.utcnow() - timedelta(days=2),
                    EndTime=datetime.utcnow(),
                    Period=86400,
                    Statistics=['Average']
                )

                datapoints = response.get('Datapoints', [])
                if datapoints:
                    size_bytes = datapoints[-1]['Average']
                    bucket_sizes[bucket_name] = size_bytes / (1024 ** 3)  # Convert to GB
                    total_bytes += size_bytes

            except Exception as e:
                print(f"Error fetching size for bucket {bucket_name}: {e}")
                continue

        total_gb = total_bytes / (1024 ** 3)
        return total_gb

    except Exception as e:
        print(f"Error listing buckets: {e}")
        return {}, 0

if __name__ == "__main__":
    main()

Fargate.py
import boto3
import csv
import openpyxl
from botocore.exceptions import ClientError

def main():
    sts_client = boto3.client('sts')
    assumed_role_object = sts_client.assume_role(
        RoleArn='arn:aws:iam::528150397796:role/528150397796-Devops-Admin-Role',
        RoleSessionName="iamautomation"
    )
    credentials = assumed_role_object['Credentials']
    
    client = boto3.client(
        'sts',
        aws_access_key_id=credentials['AccessKeyId'],
        aws_secret_access_key=credentials['SecretAccessKey'],
        aws_session_token=credentials['SessionToken']
    )

    root_acc = client.get_caller_identity().get('Account')
    print("Root Account:", root_acc)

    with open("fargate_report.csv", 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["AccountNumber", "ClusterType", "Cluster", "RunningFargateWorkloads"])
        
        assumerole_fargate(client, writer, sts_client)

def assumerole_fargate(client, writer, sts_client):
    fname = 'roles.xlsx'
    wb = openpyxl.load_workbook(fname)
    sheet = wb['Sheet1']
    
    for row in sheet['A1':'A110']:
        for cell in row:
            role_arn = cell.value
            if role_arn:
                try:
                    assumed_role_object = client.assume_role(
                        RoleArn=role_arn,
                        RoleSessionName="AssumeRoleSession1"
                    )
                    credentials = assumed_role_object['Credentials']
                    
                    ecs_client = boto3.client(
                        'ecs',
                        aws_access_key_id=credentials['AccessKeyId'],
                        aws_secret_access_key=credentials['SecretAccessKey'],
                        aws_session_token=credentials['SessionToken'],region_name="us-east-1"
                    )
                    
                    eks_client = boto3.client(
                        'eks',
                        aws_access_key_id=credentials['AccessKeyId'],
                        aws_secret_access_key=credentials['SecretAccessKey'],
                        aws_session_token=credentials['SessionToken'],region_name="us-east-1"
                    )
                    
                    sts_client1 = boto3.client(
                        'sts',
                        aws_access_key_id=credentials['AccessKeyId'],
                        aws_secret_access_key=credentials['SecretAccessKey'],
                        aws_session_token=credentials['SessionToken']
                    )
                    
                    account_num = sts_client1.get_caller_identity().get('Account')
                    print(f"Switching to Account: {account_num}")
                    
                    count_ecs_fargate_tasks(ecs_client, writer, account_num)
                    count_eks_fargate_pods(eks_client, writer, account_num)
                
                except ClientError as error:
                    print(f"Error assuming role for {role_arn}: {error}")
                    continue

def count_ecs_fargate_tasks(ecs_client, writer, account_num):
    clusters = ecs_client.list_clusters()["clusterArns"]

    for cluster in clusters:
        tasks = ecs_client.list_tasks(cluster=cluster, launchType='FARGATE', desiredStatus='RUNNING')["taskArns"]
        num_tasks = len(tasks)

        writer.writerow([account_num, "ECS", cluster, num_tasks])
        print(f"Account: {account_num}, ECS Cluster: {cluster}, Running Fargate Tasks: {num_tasks}")

def count_eks_fargate_pods(eks_client, writer, account_num):
    eks_clusters = eks_client.list_clusters()["clusters"]

    for cluster in eks_clusters:
        fargate_profiles = eks_client.list_fargate_profiles(clusterName=cluster)["fargateProfileNames"]

        if fargate_profiles:
            print(f"Account: {account_num}, EKS Cluster: {cluster}, Fargate Profiles: {fargate_profiles}")
            writer.writerow([account_num, "EKS", cluster, len(fargate_profiles)])  # Using profile count for now

if __name__ == "__main__":
    main()


iamuser.py
from multiprocessing.sharedctypes import Value
import boto3
from datetime import date
import xlwings as xw
import openpyxl
import botocore
import pandas as pd
import json
from botocore.exceptions import ClientError
def main():
  sts_client = boto3.client('sts')
  assumed_role_object = sts_client.assume_role(RoleArn='arn:aws:iam::528150397796:role/IAM_Automation',RoleSessionName="iamautomation",ExternalId="UZZCWIWGSY")
  credentials=assumed_role_object['Credentials']
  client=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
  root_acc = client.get_caller_identity().get('Account')
  print(root_acc)
  assumerole_iam(client)
def assumerole_iam(client):
  assumed_role_object = client.assume_role(RoleArn="arn:aws:iam::528150397796:role/528150397796-Devops-Admin-Role",RoleSessionName="iamautomation")
  credentials=assumed_role_object['Credentials']
  iam_client=boto3.client('iam',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
  sts_clientroot1=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
  root_acc = sts_clientroot1.get_caller_identity()
  print(root_acc)
  fname = 'roles-dev.xlsx'
  wb = openpyxl.load_workbook(fname)
  sheet = wb.get_sheet_by_name('Sheet1')
  for rowOfCellObjects in sheet['A95':'A95']:
    try:
      for cellObj in rowOfCellObjects:
        #print(cellObj.coordinate, cellObj.value)
        v1 = cellObj.value
        assumed_role_object = sts_clientroot1.assume_role(RoleArn=v1,RoleSessionName="AssumeRoleSession1")
        credentials=assumed_role_object['Credentials']
        iam_client=boto3.client('iam',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
        sts_client1=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
        account_num = sts_client1.get_caller_identity().get('Account')
        #acc_name =   boto3.client('organizations').describe_account(AccountId=account_num).get('Account').get('Name') 
        additional_tags = [{'Key': 'role_type', 'Value': 'service'}]
        additional_tags1 = [{'Key': 'role_type', 'Value': 'core'}]
        print("Switching to",account_num)
        service_roles(iam_client,additional_tags,account_num)
        core_roles(iam_client,additional_tags1,account_num)
    except botocore.exceptions.ClientError as error:
      print(error)
      continue  
def service_roles(iam_client,additional_tags,account_num):
  role_names =['rds-monitoring-role','CloudabilityRole','WizAccess-Role','Okta-Idp-cross-account-role','stacksets-exec-c1c1605357571231990026442703']
  cmr_role=account_num+"-cloud-management-Role"
  role_names.append(cmr_role)
  try:
    for role_name in role_names:
      response = iam_client.list_role_tags(RoleName=role_name)
      current_tags = response['Tags']
      # Add additional tags to the existing ones
      updated_tags = current_tags + additional_tags
      # Update tags for the IAM role
      iam_client.tag_role(RoleName=role_name, Tags=updated_tags)
      print(f"Additional tags added to IAM role {role_name} successfully.")
  except Exception as e:
    print(e)
    pass
        
def core_roles(iam_client,additional_tags1,account_num):
  role_name = account_num+"-DBE-Role"
  response = iam_client.list_role_tags(RoleName=role_name)
  current_tags = response['Tags']
  # Add additional tags to the existing ones
  updated_tags = current_tags + additional_tags1
  # Update tags for the IAM role
  iam_client.tag_role(RoleName=role_name, Tags=updated_tags)
  print(f"Additional tags added to IAM role {role_name} successfully.")

if __name__ == "__main__":
    main()

policyanalyser.py
import boto3
import csv
import openpyxl
from botocore.exceptions import ClientError
import botocore

def main():
    sts_client = boto3.client('sts')
    with open("finalreport.csv", 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["Role Name", "Policy Name", "S3 Allow", "S3 Deny", "ec2_allow", "ec2_deny","IAM Allow","IAM Deny","Cloudtrail Allow","Cloudtrail Deny","RDS Allow","RDS Deny","DynamoDB Allow","DynamoDB Deny","Administrator Access", "AccountNum"])
        root_acc = get_root_account(sts_client)
        print(root_acc)
        assumerole_iam(sts_client, writer)

def get_root_account(client):
    root_acc = client.get_caller_identity().get('Account')
    return root_acc

def assumerole_iam(sts_client, writer):
    #assumed_role_object = sts_client.assume_role(RoleArn='arn:aws:iam::528150397796:role/IAM_Automation',RoleSessionName="iamautomation",ExternalId="UZZCWIWGSY")
    #credentials=assumed_role_object['Credentials']
    #client=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
    #root_acc = client.get_caller_identity().get('Account')
    #print(root_acc)
    assumed_role_object = sts_client.assume_role(RoleArn="arn:aws:iam::528150397796:role/528150397796-Devops-Admin-Role",RoleSessionName="iamautomation")
    credentials=assumed_role_object['Credentials']
    iam_client=boto3.client('iam',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
    sts_clientroot1=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
    root_acc = sts_clientroot1.get_caller_identity()
    print(root_acc)
    fname = 'roles-dev.xlsx'
    wb = openpyxl.load_workbook(fname)
    sheet = wb.get_sheet_by_name('Sheet1')
    for rowOfCellObjects in sheet['A1':'A70']:
      try:
        for cellObj in rowOfCellObjects:
          #print(cellObj.coordinate, cellObj.value)
          v1 = cellObj.value
          assumed_role_object = sts_clientroot1.assume_role(RoleArn=v1,RoleSessionName="AssumeRoleSession1")
          credentials=assumed_role_object['Credentials']
          iam_client=boto3.client('iam',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
          sts_client1=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
          account_num = sts_client1.get_caller_identity().get('Account')
          #acc_name =   boto3.client('organizations').describe_account(AccountId=account_num).get('Account').get('Name')
          users = iam_client.list_users()["Users"]
          print("Switching to",account_num)
          iam_roles(iam_client,writer,account_num)
      except botocore.exceptions.ClientError as error: 
        print(error)
        continue
def iam_roles(iam_client, writer, account_num):
    try:
        # S3
        s3_perms_allow = ['s3:*',"s3:CreateBucket","s3:CreateAccessPoint","s3:CreateAccessGrant","s3:CreateAccessPointForObjectLambda","s3:DeleteBucket","s3:DeleteBucketWebsite","s3:DeleteAccessPoint","s3:DeleteAccessGrant","s3:DeleteObject","s3:Create*","s3:Delete*"]
        s3_perms_deny = ["s3:CreateBucket","s3:CreateAccessPoint","s3:CreateAccessGrant","s3:CreateAccessPointForObjectLambda","s3:DeleteBucket","s3:DeleteBucketWebsite","s3:DeleteAccessPoint","s3:DeleteAccessGrant","s3:DeleteObject","s3:Create*","s3:Delete*"]
        # EC2
        ec2_perms_allow = ['ec2:*',"ec2:CreateInternetGateway","ec2:DeleteInternetGateway","ec2:AttachInternetGateway","ec2:CreateNatGateway","ec2:DeleteNatGateway","ec2:CreateVpcPeeringConnection","ec2:CreateNetworkAcl","ec2:DeleteNetworkAcl","ec2:DisassociateNatGatewayAddress","ec2:DisassociateRouteTable","ec2:DisassociateSubnetCidrBlock","ec2:DisassociateVpcCidrBlock","ec2:AuthorizeClientVpnIngress","ec2:AcceptTransitGatewayPeeringAttachment","ec2:AcceptTransitGatewayVpcAttachment","ec2:ModifySubnetAttribute","ec2:ModifyTransitGatewayVpcAttachment","ec2:CreateSubnet","ec2:AcceptVpcPeeringConnection","ec2:DeleteVpcPeeringConnection","ec2:CreateVpc","ec2:DeleteVpc","ec2:ModifyVpcTenancy","ec2:CreateFlowLogs","ec2:DeleteFlowLogs","ec2:AttachVpnGateway","ec2:CreateVpnGateway","ec2:DeleteVpnGateway","ec2:DisableVgwRoutePropagation","ec2:EnableVgwRoutePropagation","ec2:CreateVpnConnectionRoute","ec2:DeleteVpnConnection","ec2:DeleteVpnConnectionRoute","ec2:ModifyVpnConnection","ec2:CreateCustomerGateway","ec2:DeleteCustomerGateway","ec2:CreateRouteTable","ec2:AssociateRouteTable","ec2:CreateRoute","ec2:DeleteRouteTable","ec2:ModifyVpcAttribute","ec2:ReplaceRoute","ec2:DeleteRoute","ec2:CreateTransitGateway","ec2:DeleteTransitGatewayRouteTable","ec2:CreateTransitGatewayRouteTable","ec2:ReplaceTransitGatewayRoute"]
        ec2_perms_deny = ["ec2:CreateInternetGateway","ec2:DeleteInternetGateway","ec2:AttachInternetGateway","ec2:CreateNatGateway","ec2:DeleteNatGateway","ec2:CreateVpcPeeringConnection","ec2:CreateNetworkAcl","ec2:DeleteNetworkAcl","ec2:DisassociateNatGatewayAddress","ec2:DisassociateRouteTable","ec2:DisassociateSubnetCidrBlock","ec2:DisassociateVpcCidrBlock","ec2:AuthorizeClientVpnIngress","ec2:AcceptTransitGatewayPeeringAttachment","ec2:AcceptTransitGatewayVpcAttachment","ec2:ModifySubnetAttribute","ec2:ModifyTransitGatewayVpcAttachment","ec2:CreateSubnet","ec2:AcceptVpcPeeringConnection","ec2:DeleteVpcPeeringConnection","ec2:CreateVpc","ec2:DeleteVpc","ec2:ModifyVpcTenancy","ec2:CreateFlowLogs","ec2:DeleteFlowLogs","ec2:AttachVpnGateway","ec2:CreateVpnGateway","ec2:DeleteVpnGateway","ec2:DisableVgwRoutePropagation","ec2:EnableVgwRoutePropagation","ec2:CreateVpnConnectionRoute","ec2:DeleteVpnConnection","ec2:DeleteVpnConnectionRoute","ec2:ModifyVpnConnection","ec2:CreateCustomerGateway","ec2:DeleteCustomerGateway","ec2:CreateRouteTable","ec2:AssociateRouteTable","ec2:CreateRoute","ec2:DeleteRouteTable","ec2:ModifyVpcAttribute","ec2:ReplaceRoute","ec2:DeleteRoute","ec2:CreateTransitGateway","ec2:DeleteTransitGatewayRouteTable","ec2:CreateTransitGatewayRouteTable","ec2:ReplaceTransitGatewayRoute"]
        #IAM
        iam_perms_allow = ['iam:*',"iam:CreateAccessKey","iam:CreateRole","iam:CreateUser","iam:DeleteAccessKey","iam:DeleteGroup","iam:DeleteRole","iam:DeleteUser"]
        iam_perms_deny = ["iam:CreateUser","iam:DeleteUser","iam:UpdateUser","iam:CreateRole","iam:DeleteRole","iam:PassRole","iam:UpdateRole"]
        # Cloudtrail
        cloudtrail_perms_allow =["cloudtrail:*","cloudtrail:CreateTrail","cloudtrail:DeleteTrail","cloudtrail:UpdateTrail"]
        cloudtrail_perms_deny = ["cloudtrail:CreateTrail","cloudtrail:DeleteTrail","cloudtrail:UpdateTrail"]
        # RDS
        rds_perms_allow = ['rds:*',"rds:CreateDBInstance","rds:CreateDBClusterSnapshot","rds:CreateDBCluster","rds:CreateDBSecurityGroup","rds:DeleteDBCluster","rds:DeleteDBInstance","rds:DeleteDBSecurityGroup","elasticfilesystem:Create*","elasticfilesystem:Delete*"]
        rds_perms_deny = ["rds:CreateDBInstance","rds:CreateDBClusterSnapshot","rds:CreateDBCluster","rds:CreateDBSecurityGroup","rds:DeleteDBCluster","rds:DeleteDBInstance","rds:DeleteDBSecurityGroup","elasticfilesystem:Create*","elasticfilesystem:Delete*"]
        # DynamoDB
        dynamodb_perms_allow = ['dynamodb:*',"dynamodb:CreateBackup","dynamodb:CreateGlobalTable","dynamodb:CreateTable","dynamodb:CreateTableReplica","dynamodb:DeleteBackup","dynamodb:DeleteItem","dynamodb:DeleteTable","dynamodb:DeleteTableReplica","dynamodb:DeleteResourcePolicy","dynamodb:PartiQLDelete","dynamodb:Create*","dynamodb:Delete*"]
        dynamodb_perms_deny = ["dynamodb:CreateBackup","dynamodb:CreateGlobalTable","dynamodb:CreateTable","dynamodb:CreateTableReplica","dynamodb:DeleteBackup","dynamodb:DeleteItem","dynamodb:DeleteTable","dynamodb:DeleteTableReplica","dynamodb:DeleteResourcePolicy","dynamodb:PartiQLDelete","dynamodb:Create*","dynamodb:Delete*"]
        # Admin
        admin_access = ['*',"all"]
        
        paginator = iam_client.get_paginator("list_roles")
        for response in paginator.paginate(PaginationConfig={'MaxItems': 1000}):
            for role in response['Roles']:
                role_name = role["RoleName"]
                role_arn = role['Arn']
                attached_policies = iam_client.list_attached_role_policies(RoleName=role['RoleName'])['AttachedPolicies']
                policies_list = []
                s3_allow_list = []
                s3_deny_list = []
                ec2_allow_list = []
                ec2_deny_list = []
                iam_allow_list = []
                iam_deny_list = []
                cloudtrail_allow_list = []
                cloudtrail_deny_list = []
                rds_allow_list = []
                rds_deny_list = []
                dynamodb_allow_list = []
                dynamodb_deny_list = []
                admin_list = []
                
                for policy in attached_policies:
                    #policy_version = iam_client.get_policy_version(PolicyArn=policy['PolicyArn'], VersionId=policy['DefaultVersionId'])['PolicyVersion']['Document']
                    policy_document = iam_client.get_policy(PolicyArn=policy['PolicyArn'])['Policy']['DefaultVersionId']
                    print(policy)
                    policy_version = iam_client.get_policy_version(PolicyArn=policy['PolicyArn'],VersionId=policy_document)['PolicyVersion']['Document']
                    process_policy_statement(policy_version, s3_perms_allow, s3_allow_list, s3_perms_deny, s3_deny_list, ec2_perms_allow, ec2_allow_list, ec2_perms_deny, ec2_deny_list,iam_perms_allow,iam_allow_list, iam_perms_deny, iam_deny_list, cloudtrail_perms_allow, cloudtrail_allow_list, cloudtrail_perms_deny, cloudtrail_deny_list, rds_perms_allow, rds_allow_list, rds_perms_deny, rds_deny_list, dynamodb_perms_allow, dynamodb_allow_list, dynamodb_perms_deny, dynamodb_deny_list, admin_access, admin_list)
                    policies_list.append(policy['PolicyName'])
                    
                inline_policies = iam_client.list_role_policies(RoleName=role['RoleName'])['PolicyNames']
                for policy_name in inline_policies:
                    policy_version = iam_client.get_role_policy(RoleName=role['RoleName'], PolicyName=policy_name)['PolicyDocument']
                    process_policy_statement(policy_version, s3_perms_allow, s3_allow_list, s3_perms_deny, s3_deny_list, ec2_perms_allow, ec2_allow_list, ec2_perms_deny, ec2_deny_list,iam_perms_allow,iam_allow_list, iam_perms_deny, iam_deny_list, cloudtrail_perms_allow, cloudtrail_allow_list, cloudtrail_perms_deny, cloudtrail_deny_list, rds_perms_allow, rds_allow_list, rds_perms_deny, rds_deny_list, dynamodb_perms_allow, dynamodb_allow_list, dynamodb_perms_deny, dynamodb_deny_list, admin_access, admin_list)
                    policies_list.append(policy_name)
                writer.writerow([role_name, ','.join(map(str,policies_list)), ','.join(map(str,set(s3_allow_list))), ','.join(map(str,set(s3_deny_list))), ','.join(map(str,set(ec2_allow_list))), ','.join(map(str,set(ec2_deny_list))),','.join(map(str,set(iam_allow_list))),','.join(map(str,set(iam_deny_list))),','.join(map(str,set(cloudtrail_allow_list))), ','.join(map(str,set(cloudtrail_deny_list))), ','.join(map(str,set(rds_allow_list))), ','.join(map(str,set(rds_deny_list))),','.join(map(str,set(dynamodb_allow_list))), ','.join(map(str,set(dynamodb_deny_list))),','.join(map(str,set(admin_list))), account_num,role_arn])
                print(role_name, "---->", policies_list, s3_allow_list, s3_deny_list, ec2_allow_list, ec2_deny_list, admin_access, account_num)
    except Exception as e:
        print("Error:", e)

def process_policy_statement(policy_version, s3_perms_allow, s3_allow_list, s3_perms_deny, s3_deny_list, ec2_perms_allow, ec2_allow_list, ec2_perms_deny, ec2_deny_list,iam_perms_allow,iam_allow_list, iam_perms_deny, iam_deny_list, cloudtrail_perms_allow, cloudtrail_allow_list, cloudtrail_perms_deny, cloudtrail_deny_list, rds_perms_allow, rds_allow_list, rds_perms_deny, rds_deny_list, dynamodb_perms_allow, dynamodb_allow_list, dynamodb_perms_deny, dynamodb_deny_list, admin_access, admin_list):
    try:
      for idx, statement in enumerate(policy_version.get('Statement', []), start=1):
        effect = statement.get('Effect', 'N/A')
        actions = statement.get('Action', 'N/A')
        resources = statement.get('Resource', 'N/A')
        conditions = statement.get('Condition', 'N/A')
        if resources !="*":
            print("Resources are there:",resources)
        elif conditions is None:
            if effect == 'Allow':
                for s3_perm_allow in s3_perms_allow:
                    if s3_perm_allow in actions:
                        s3_allow_list.append(s3_perm_allow)
                for ec2_perm_allow in ec2_perms_allow:
                    if ec2_perm_allow in actions:
                        ec2_allow_list.append(ec2_perm_allow)
                for iam_perm_allow in iam_perms_allow:
                    if iam_perm_allow in actions:
                        iam_allow_list.append(iam_perm_allow)
                for cloudtrail_perm_allow in cloudtrail_perms_allow:
                    if cloudtrail_perm_allow in actions:
                        cloudtrail_allow_list.append(cloudtrail_perm_allow)
                for rds_perm_allow in rds_perms_allow:
                    if rds_perm_allow in actions:
                        rds_allow_list.append(rds_perm_allow)
                for dynamodb_perm_allow in dynamodb_perms_allow:
                    if dynamodb_perm_allow in actions:
                        dynamodb_allow_list.append(dynamodb_perm_allow)
                for admin_acc in admin_access:
                    if admin_acc in actions:
                        admin_list.append("admin_accees")     
            if effect == 'Deny':
                for s3_perm_deny in s3_perms_deny:
                    if s3_perm_deny in actions:
                        s3_deny_list.append(s3_perm_deny)
                for ec2_perm_deny in ec2_perms_deny:
                    if ec2_perm_deny in actions:
                        ec2_deny_list.append(ec2_perm_deny)
                for iam_perm_deny in iam_perms_deny:
                    if iam_perm_deny in actions:
                        iam_deny_list.append(iam_perm_deny)
                for cloudtrail_perm_deny in cloudtrail_perms_deny:
                    if cloudtrail_perm_deny in actions:
                        cloudtrail_deny_list.append(cloudtrail_perm_deny)
                for rds_perm_deny in rds_perms_deny:
                    if rds_perm_deny in actions:
                        rds_deny_list.append(rds_perm_deny)
                for dynamodb_perm_deny in dynamodb_perms_deny:
                    if dynamodb_perm_deny in actions:
                        dynamodb_deny_list.append(dynamodb_perm_deny)        
    except Exception as e1:
      print(e1)
      pass
if __name__ == "__main__":
    main()


Role_tagging.py
from multiprocessing.sharedctypes import Value
import boto3
from datetime import date
import xlwings as xw
import openpyxl
import botocore
import pandas as pd
import json
from botocore.exceptions import ClientError
def main():
  sts_client = boto3.client('sts')
  assumed_role_object = sts_client.assume_role(RoleArn='arn:aws:iam::528150397796:role/IAM_Automation',RoleSessionName="iamautomation",ExternalId="UZZCWIWGSY")
  credentials=assumed_role_object['Credentials']
  client=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
  root_acc = client.get_caller_identity().get('Account')
  print(root_acc)
  assumerole_iam(client)
def assumerole_iam(client):
  assumed_role_object = client.assume_role(RoleArn="arn:aws:iam::528150397796:role/528150397796-Devops-Admin-Role",RoleSessionName="iamautomation")
  credentials=assumed_role_object['Credentials']
  iam_client=boto3.client('iam',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
  sts_clientroot1=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
  root_acc = sts_clientroot1.get_caller_identity()
  print(root_acc)
  fname = 'Prodaccounts.xlsx'
  wb = openpyxl.load_workbook(fname)
  sheet = wb.get_sheet_by_name('Sheet1')
  for rowOfCellObjects in sheet['A2':'A2']:
    try:
      for cellObj in rowOfCellObjects:
        #print(cellObj.coordinate, cellObj.value)
        v1 = cellObj.value
        assumed_role_object = sts_clientroot1.assume_role(RoleArn=v1,RoleSessionName="AssumeRoleSession1")
        credentials=assumed_role_object['Credentials']
        iam_client=boto3.client('iam',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
        sts_client1=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
        account_num = sts_client1.get_caller_identity().get('Account')
        #acc_name =   boto3.client('organizations').describe_account(AccountId=account_num).get('Account').get('Name') 
        additional_tags = [{'Key': 'role_type', 'Value': 'service'}]
        additional_tags1 = [{'Key': 'role_type', 'Value': 'core'}]
        f = open("issues-service.csv",'w')
        f1 = open("issues-core.csv",'w')
        print("Switching to",account_num)
        service_roles(iam_client,additional_tags,account_num,f)
        core_roles(iam_client,additional_tags1,account_num,f1)
    except botocore.exceptions.ClientError as error:
      print(error)
      continue  
def service_roles(iam_client,additional_tags,account_num,f):
  role_names =['rds-monitoring-role','CloudabilityRole','WizAccess-Role','Okta-Idp-cross-account-role','stacksets-exec-c1c1605357571231990026442703']
  cmr_role=account_num+"-cloud-management-Role"
  role_names.append(cmr_role)
  try:
    for role_name in role_names:
      response = iam_client.list_role_tags(RoleName=role_name)
      current_tags = response['Tags']
      # Add additional tags to the existing ones
      updated_tags = current_tags + additional_tags
      # Update tags for the IAM role
      iam_client.tag_role(RoleName=role_name, Tags=updated_tags)
      print(f"Additional tags added to IAM role {role_name} successfully.")
  except (botocore.errorfactory.InvalidInputException,boto3.exceptions.NotFoundException) as e:
    print(e)
    f.write(role_name+","+account_num+","+str(e)+"\n")
    pass
        
def core_roles(iam_client,additional_tags1,account_num,f1):
  try:
    role_name = account_num+"-DBE-Role"
    response = iam_client.list_role_tags(RoleName=role_name)
    current_tags = response['Tags']
    # Add additional tags to the existing ones
    updated_tags = current_tags + additional_tags1
    # Update tags for the IAM role
    iam_client.tag_role(RoleName=role_name, Tags=updated_tags)
    print(f"Additional tags added to IAM role {role_name} successfully.")
  except (botocore.errorfactory.InvalidInputException,boto3.exceptions.NotFoundException) as e:
    print(e)
    f1.write(role_name+","+account_num+","+str(e)+"\n")
    pass
if __name__ == "__main__":
    main()


rolecompliance.py
import boto3
import csv
import openpyxl
from botocore.exceptions import ClientError
import botocore
import json

def main():
    sts_client = boto3.client('sts')
    with open("finalreport.csv", 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["Role Name", "Policy Name", "S3 Allow", "S3 Deny", "ec2_allow", "ec2_deny","IAM Allow","IAM Deny","Cloudtrail Allow","Cloudtrail Deny","RDS Allow","RDS Deny","DynamoDB Allow","DynamoDB Deny","Administrator Access", "AccountNum","Role ARN","Compliance/NonCompliance"])
        root_acc = get_root_account(sts_client)
        print(root_acc)
        assumerole_iam(sts_client, writer)

def get_root_account(client):
    root_acc = client.get_caller_identity().get('Account')
    return root_acc

def assumerole_iam(sts_client, writer):
    #assumed_role_object = sts_client.assume_role(RoleArn='arn:aws:iam::528150397796:role/IAM_Automation',RoleSessionName="iamautomation",ExternalId="UZZCWIWGSY")
    #credentials=assumed_role_object['Credentials']
    #client=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
    #root_acc = client.get_caller_identity().get('Account')
    #print(root_acc)
    assumed_role_object = sts_client.assume_role(RoleArn="arn:aws:iam::528150397796:role/528150397796-Devops-Admin-Role",RoleSessionName="iamautomation")
    credentials=assumed_role_object['Credentials']
    iam_client=boto3.client('iam',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
    sts_clientroot1=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
    root_acc = sts_clientroot1.get_caller_identity()
    print(root_acc)
    fname = 'roles.xlsx'
    wb = openpyxl.load_workbook(fname)
    sheet = wb.get_sheet_by_name('Sheet1')
    for rowOfCellObjects in sheet['A56':'A110']:
      try:
        for cellObj in rowOfCellObjects:
          #print(cellObj.coordinate, cellObj.value)
          v1 = cellObj.value
          assumed_role_object = sts_clientroot1.assume_role(RoleArn=v1,RoleSessionName="AssumeRoleSession1")
          credentials=assumed_role_object['Credentials']
          iam_client=boto3.client('iam',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
          sts_client1=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
          account_num = sts_client1.get_caller_identity().get('Account')
          #acc_name =   boto3.client('organizations').describe_account(AccountId=account_num).get('Account').get('Name')
          users = iam_client.list_users()["Users"]
          print("Switching to",account_num)
          iam_roles(iam_client,writer,account_num)
      except botocore.exceptions.ClientError as error: 
        print(error)
        continue
def iam_roles(iam_client, writer, account_num):
    try:
        # S3
        s3_perms_allow = ['s3:*',"s3:CreateBucket","s3:CreateAccessPoint","s3:CreateAccessGrant","s3:CreateAccessPointForObjectLambda","s3:DeleteBucket","s3:DeleteBucketWebsite","s3:DeleteAccessPoint","s3:DeleteAccessGrant","s3:DeleteObject","s3:Create*","s3:Delete*"]
        s3_perms_deny = ["s3:CreateBucket","s3:CreateAccessPoint","s3:CreateAccessGrant","s3:CreateAccessPointForObjectLambda","s3:DeleteBucket","s3:DeleteBucketWebsite","s3:DeleteAccessPoint","s3:DeleteAccessGrant","s3:DeleteObject","s3:Create*","s3:Delete*"]
        # EC2
        ec2_perms_allow = ['ec2:*',"ec2:CreateInternetGateway","ec2:DeleteInternetGateway","ec2:AttachInternetGateway","ec2:CreateNatGateway","ec2:DeleteNatGateway","ec2:CreateVpcPeeringConnection","ec2:CreateNetworkAcl","ec2:DeleteNetworkAcl","ec2:DisassociateNatGatewayAddress","ec2:DisassociateRouteTable","ec2:DisassociateSubnetCidrBlock","ec2:DisassociateVpcCidrBlock","ec2:AuthorizeClientVpnIngress","ec2:AcceptTransitGatewayPeeringAttachment","ec2:AcceptTransitGatewayVpcAttachment","ec2:ModifySubnetAttribute","ec2:ModifyTransitGatewayVpcAttachment","ec2:CreateSubnet","ec2:AcceptVpcPeeringConnection","ec2:DeleteVpcPeeringConnection","ec2:CreateVpc","ec2:DeleteVpc","ec2:ModifyVpcTenancy","ec2:CreateFlowLogs","ec2:DeleteFlowLogs","ec2:AttachVpnGateway","ec2:CreateVpnGateway","ec2:DeleteVpnGateway","ec2:DisableVgwRoutePropagation","ec2:EnableVgwRoutePropagation","ec2:CreateVpnConnectionRoute","ec2:DeleteVpnConnection","ec2:DeleteVpnConnectionRoute","ec2:ModifyVpnConnection","ec2:CreateCustomerGateway","ec2:DeleteCustomerGateway","ec2:CreateRouteTable","ec2:AssociateRouteTable","ec2:CreateRoute","ec2:DeleteRouteTable","ec2:ModifyVpcAttribute","ec2:ReplaceRoute","ec2:DeleteRoute","ec2:CreateTransitGateway","ec2:DeleteTransitGatewayRouteTable","ec2:CreateTransitGatewayRouteTable","ec2:ReplaceTransitGatewayRoute"]
        ec2_perms_deny = ["ec2:CreateInternetGateway","ec2:DeleteInternetGateway","ec2:AttachInternetGateway","ec2:CreateNatGateway","ec2:DeleteNatGateway","ec2:CreateVpcPeeringConnection","ec2:CreateNetworkAcl","ec2:DeleteNetworkAcl","ec2:DisassociateNatGatewayAddress","ec2:DisassociateRouteTable","ec2:DisassociateSubnetCidrBlock","ec2:DisassociateVpcCidrBlock","ec2:AuthorizeClientVpnIngress","ec2:AcceptTransitGatewayPeeringAttachment","ec2:AcceptTransitGatewayVpcAttachment","ec2:ModifySubnetAttribute","ec2:ModifyTransitGatewayVpcAttachment","ec2:CreateSubnet","ec2:AcceptVpcPeeringConnection","ec2:DeleteVpcPeeringConnection","ec2:CreateVpc","ec2:DeleteVpc","ec2:ModifyVpcTenancy","ec2:CreateFlowLogs","ec2:DeleteFlowLogs","ec2:AttachVpnGateway","ec2:CreateVpnGateway","ec2:DeleteVpnGateway","ec2:DisableVgwRoutePropagation","ec2:EnableVgwRoutePropagation","ec2:CreateVpnConnectionRoute","ec2:DeleteVpnConnection","ec2:DeleteVpnConnectionRoute","ec2:ModifyVpnConnection","ec2:CreateCustomerGateway","ec2:DeleteCustomerGateway","ec2:CreateRouteTable","ec2:AssociateRouteTable","ec2:CreateRoute","ec2:DeleteRouteTable","ec2:ModifyVpcAttribute","ec2:ReplaceRoute","ec2:DeleteRoute","ec2:CreateTransitGateway","ec2:DeleteTransitGatewayRouteTable","ec2:CreateTransitGatewayRouteTable","ec2:ReplaceTransitGatewayRoute"]
        #IAM
        iam_perms_allow = ['iam:*',"iam:CreateAccessKey","iam:CreateRole","iam:CreateUser","iam:DeleteAccessKey","iam:DeleteGroup","iam:DeleteRole","iam:DeleteUser"]
        iam_perms_deny = ["iam:CreateUser","iam:DeleteUser","iam:UpdateUser","iam:CreateRole","iam:DeleteRole","iam:PassRole","iam:UpdateRole"]
        # Cloudtrail
        cloudtrail_perms_allow =["cloudtrail:*","cloudtrail:CreateTrail","cloudtrail:DeleteTrail","cloudtrail:UpdateTrail"]
        cloudtrail_perms_deny = ["cloudtrail:CreateTrail","cloudtrail:DeleteTrail","cloudtrail:UpdateTrail"]
        # RDS
        rds_perms_allow = ['rds:*',"rds:CreateDBInstance","rds:CreateDBClusterSnapshot","rds:CreateDBCluster","rds:CreateDBSecurityGroup","rds:DeleteDBCluster","rds:DeleteDBInstance","rds:DeleteDBSecurityGroup","elasticfilesystem:Create*","elasticfilesystem:Delete*"]
        rds_perms_deny = ["rds:CreateDBInstance","rds:CreateDBClusterSnapshot","rds:CreateDBCluster","rds:CreateDBSecurityGroup","rds:DeleteDBCluster","rds:DeleteDBInstance","rds:DeleteDBSecurityGroup","elasticfilesystem:Create*","elasticfilesystem:Delete*"]
        # DynamoDB
        dynamodb_perms_allow = ['dynamodb:*',"dynamodb:CreateBackup","dynamodb:CreateGlobalTable","dynamodb:CreateTable","dynamodb:CreateTableReplica","dynamodb:DeleteBackup","dynamodb:DeleteItem","dynamodb:DeleteTable","dynamodb:DeleteTableReplica","dynamodb:DeleteResourcePolicy","dynamodb:PartiQLDelete","dynamodb:Create*","dynamodb:Delete*"]
        dynamodb_perms_deny = ["dynamodb:CreateBackup","dynamodb:CreateGlobalTable","dynamodb:CreateTable","dynamodb:CreateTableReplica","dynamodb:DeleteBackup","dynamodb:DeleteItem","dynamodb:DeleteTable","dynamodb:DeleteTableReplica","dynamodb:DeleteResourcePolicy","dynamodb:PartiQLDelete","dynamodb:Create*","dynamodb:Delete*"]
        # Admin
        admin_access = ['*']
        role_names =['rds-monitoring-role','CloudabilityRole','WizAccess-Role','Okta-Idp-cross-account-role','stacksets-exec-c1c1b0535f75d712e3199a0026442703','CloudWatch-CrossAccountSharingRole']
        CMRrole=account_num+"-cloud-management-Role"
        role_names.append(CMRrole)
        Devopsrole = account_num+"-Devops-Admin-Role"
        role_names.append(Devopsrole)
        Readonlyrole = account_num+"-Readonly-Admin-Role"
        role_names.append(Readonlyrole)
        Financialrole = account_num+"-Financial-Analyst-Role"
        role_names.append(Financialrole)
        DBErole = account_num+"-DBE-Role"
        role_names.append(DBErole)
        #print(role_names)
        okta ="Okta"
        AppCore = "AppCore"
        Application = "Application"
        AWSServiceroles="AWSServiceRoleFor"
        paginator = iam_client.get_paginator("list_roles")
        for response in paginator.paginate(PaginationConfig={'MaxItems': 1000}):
            for role in response['Roles']:
                role_name = role["RoleName"]
                assume_doc = role['AssumeRolePolicyDocument']
                assume_string = json.dumps(assume_doc)
                if okta in assume_string:
                  if role_name in role_names:
                    print("Role is core/service",role_name)
                  else:
                    role_arn = role['Arn']
                    attached_policies = iam_client.list_attached_role_policies(RoleName=role['RoleName'])['AttachedPolicies']
                    policies_list = []
                    s3_allow_list = []
                    s3_deny_list = []
                    ec2_allow_list = []
                    ec2_deny_list = []
                    iam_allow_list = []
                    iam_deny_list = []
                    cloudtrail_allow_list = []
                    cloudtrail_deny_list = []
                    rds_allow_list = []
                    rds_deny_list = []
                    dynamodb_allow_list = []
                    dynamodb_deny_list = []
                    admin_list = []
                    compliance = " Compliance Role"
                    noncompliance = "NonCompliance Role"

                    for policy in attached_policies:
                        #policy_version = iam_client.get_policy_version(PolicyArn=policy['PolicyArn'], VersionId=policy['DefaultVersionId'])['PolicyVersion']['Document']
                        policy_document = iam_client.get_policy(PolicyArn=policy['PolicyArn'])['Policy']['DefaultVersionId']
                        #print(policy)
                        policy_version = iam_client.get_policy_version(PolicyArn=policy['PolicyArn'],VersionId=policy_document)['PolicyVersion']['Document']
                        process_policy_statement(policy_version, s3_perms_allow, s3_allow_list, s3_perms_deny, s3_deny_list, ec2_perms_allow, ec2_allow_list, ec2_perms_deny, ec2_deny_list,iam_perms_allow,iam_allow_list, iam_perms_deny, iam_deny_list, cloudtrail_perms_allow, cloudtrail_allow_list, cloudtrail_perms_deny, cloudtrail_deny_list, rds_perms_allow, rds_allow_list, rds_perms_deny, rds_deny_list, dynamodb_perms_allow, dynamodb_allow_list, dynamodb_perms_deny, dynamodb_deny_list, admin_access, admin_list)
                        policies_list.append(policy['PolicyName'])
                        if policy['PolicyName'] == 'AdministratorAccess':
                          admin_list.append("admin_access")

                    inline_policies = iam_client.list_role_policies(RoleName=role['RoleName'])['PolicyNames']
                    for policy_name in inline_policies:
                        policy_version = iam_client.get_role_policy(RoleName=role['RoleName'], PolicyName=policy_name)['PolicyDocument']
                        process_policy_statement(policy_version, s3_perms_allow, s3_allow_list, s3_perms_deny, s3_deny_list, ec2_perms_allow, ec2_allow_list, ec2_perms_deny, ec2_deny_list,iam_perms_allow,iam_allow_list, iam_perms_deny, iam_deny_list, cloudtrail_perms_allow, cloudtrail_allow_list, cloudtrail_perms_deny, cloudtrail_deny_list, rds_perms_allow, rds_allow_list, rds_perms_deny, rds_deny_list, dynamodb_perms_allow, dynamodb_allow_list, dynamodb_perms_deny, dynamodb_deny_list, admin_access, admin_list)
                        policies_list.append(policy_name)
                    total_allow = s3_allow_list + ec2_allow_list + iam_allow_list + cloudtrail_allow_list + rds_allow_list + dynamodb_allow_list + admin_list

                    if (s3_allow_list and not s3_deny_list) or (ec2_allow_list and not ec2_deny_list) or (iam_allow_list and not iam_deny_list) or (cloudtrail_allow_list and not cloudtrail_deny_list) or (rds_allow_list and not rds_deny_list) or (dynamodb_allow_list and not dynamodb_deny_list) or admin_list:
                      writer.writerow([role_name, ','.join(map(str,policies_list)), ','.join(map(str,set(s3_allow_list))), ','.join(map(str,set(s3_deny_list))), ','.join(map(str,set(ec2_allow_list))), ','.join(map(str,set(ec2_deny_list))),','.join(map(str,set(iam_allow_list))),','.join(map(str,set(iam_deny_list))),','.join(map(str,set(cloudtrail_allow_list))), ','.join(map(str,set(cloudtrail_deny_list))), ','.join(map(str,set(rds_allow_list))), ','.join(map(str,set(rds_deny_list))),','.join(map(str,set(dynamodb_allow_list))), ','.join(map(str,set(dynamodb_deny_list))),','.join(map(str,set(admin_list))), account_num,role_arn,AppCore, noncompliance])
                      print(role_name, f"{AppCore}noncompliance ---->", policies_list, s3_allow_list, s3_deny_list, ec2_allow_list, ec2_deny_list, admin_list, account_num)

                    else:
                      writer.writerow([role_name, ','.join(map(str,policies_list)), ','.join(map(str,set(s3_allow_list))), ','.join(map(str,set(s3_deny_list))), ','.join(map(str,set(ec2_allow_list))), ','.join(map(str,set(ec2_deny_list))),','.join(map(str,set(iam_allow_list))),','.join(map(str,set(iam_deny_list))),','.join(map(str,set(cloudtrail_allow_list))), ','.join(map(str,set(cloudtrail_deny_list))), ','.join(map(str,set(rds_allow_list))), ','.join(map(str,set(rds_deny_list))),','.join(map(str,set(dynamodb_allow_list))), ','.join(map(str,set(dynamodb_deny_list))),','.join(map(str,set(admin_list))), account_num,role_arn, AppCore, compliance])
                      print(role_name, f"{AppCore}compliance ---->", policies_list, s3_allow_list, s3_deny_list, ec2_allow_list, ec2_deny_list, admin_list, account_num)
                elif AWSServiceroles in role_name:
                    if role_name in role_names:
                        print("Role is core/service",role_name)
                    else:
                        role_arn = role['Arn']
                        attached_policies = iam_client.list_attached_role_policies(RoleName=role['RoleName'])['AttachedPolicies']
                        policies_list = []
                        s3_allow_list = []
                        s3_deny_list = []
                        ec2_allow_list = []
                        ec2_deny_list = []
                        iam_allow_list = []
                        iam_deny_list = []
                        cloudtrail_allow_list = []
                        cloudtrail_deny_list = []
                        rds_allow_list = []
                        rds_deny_list = []
                        dynamodb_allow_list = []
                        dynamodb_deny_list = []
                        admin_list = []
                        compliance = " Compliance Role"
                        noncompliance = "NonCompliance Role"

                        for policy in attached_policies:
                            #policy_version = iam_client.get_policy_version(PolicyArn=policy['PolicyArn'], VersionId=policy['DefaultVersionId'])['PolicyVersion']['Document']
                            policy_document = iam_client.get_policy(PolicyArn=policy['PolicyArn'])['Policy']['DefaultVersionId']
                            #print(policy)
                            policy_version = iam_client.get_policy_version(PolicyArn=policy['PolicyArn'],VersionId=policy_document)['PolicyVersion']['Document']
                            process_policy_statement(policy_version, s3_perms_allow, s3_allow_list, s3_perms_deny, s3_deny_list, ec2_perms_allow, ec2_allow_list, ec2_perms_deny, ec2_deny_list,iam_perms_allow,iam_allow_list, iam_perms_deny, iam_deny_list, cloudtrail_perms_allow, cloudtrail_allow_list, cloudtrail_perms_deny, cloudtrail_deny_list, rds_perms_allow, rds_allow_list, rds_perms_deny, rds_deny_list, dynamodb_perms_allow, dynamodb_allow_list, dynamodb_perms_deny, dynamodb_deny_list, admin_access, admin_list)
                            policies_list.append(policy['PolicyName'])
                            if policy['PolicyName'] == 'AdministratorAccess':
                                admin_list.append("admin_access")

                        inline_policies = iam_client.list_role_policies(RoleName=role['RoleName'])['PolicyNames']
                        for policy_name in inline_policies:
                            policy_version = iam_client.get_role_policy(RoleName=role['RoleName'], PolicyName=policy_name)['PolicyDocument']
                            process_policy_statement(policy_version, s3_perms_allow, s3_allow_list, s3_perms_deny, s3_deny_list, ec2_perms_allow, ec2_allow_list, ec2_perms_deny, ec2_deny_list,iam_perms_allow,iam_allow_list, iam_perms_deny, iam_deny_list, cloudtrail_perms_allow, cloudtrail_allow_list, cloudtrail_perms_deny, cloudtrail_deny_list, rds_perms_allow, rds_allow_list, rds_perms_deny, rds_deny_list, dynamodb_perms_allow, dynamodb_allow_list, dynamodb_perms_deny, dynamodb_deny_list, admin_access, admin_list)
                            policies_list.append(policy_name)
                        total_allow = s3_allow_list + ec2_allow_list + iam_allow_list + cloudtrail_allow_list + rds_allow_list + dynamodb_allow_list + admin_list

                        if (s3_allow_list and not s3_deny_list) or (ec2_allow_list and not ec2_deny_list) or (iam_allow_list and not iam_deny_list) or (cloudtrail_allow_list and not cloudtrail_deny_list) or (rds_allow_list and not rds_deny_list) or (dynamodb_allow_list and not dynamodb_deny_list) or admin_list:
                            writer.writerow([role_name, ','.join(map(str,policies_list)), ','.join(map(str,set(s3_allow_list))), ','.join(map(str,set(s3_deny_list))), ','.join(map(str,set(ec2_allow_list))), ','.join(map(str,set(ec2_deny_list))),','.join(map(str,set(iam_allow_list))),','.join(map(str,set(iam_deny_list))),','.join(map(str,set(cloudtrail_allow_list))), ','.join(map(str,set(cloudtrail_deny_list))), ','.join(map(str,set(rds_allow_list))), ','.join(map(str,set(rds_deny_list))),','.join(map(str,set(dynamodb_allow_list))), ','.join(map(str,set(dynamodb_deny_list))),','.join(map(str,set(admin_list))), account_num,role_arn, AWSServiceroles, noncompliance])
                            print(role_name, f"{AWSServiceroles}noncompliance ---->", policies_list, s3_allow_list, s3_deny_list, ec2_allow_list, ec2_deny_list, admin_list, account_num)

                        else:
                            writer.writerow([role_name, ','.join(map(str,policies_list)), ','.join(map(str,set(s3_allow_list))), ','.join(map(str,set(s3_deny_list))), ','.join(map(str,set(ec2_allow_list))), ','.join(map(str,set(ec2_deny_list))),','.join(map(str,set(iam_allow_list))),','.join(map(str,set(iam_deny_list))),','.join(map(str,set(cloudtrail_allow_list))), ','.join(map(str,set(cloudtrail_deny_list))), ','.join(map(str,set(rds_allow_list))), ','.join(map(str,set(rds_deny_list))),','.join(map(str,set(dynamodb_allow_list))), ','.join(map(str,set(dynamodb_deny_list))),','.join(map(str,set(admin_list))), account_num,role_arn, AWSServiceroles,compliance])
                            print(role_name, f"{AWSServiceroles}compliance ---->", policies_list, s3_allow_list, s3_deny_list, ec2_allow_list, ec2_deny_list, admin_list, account_num)
                else:
                    if role_name in role_names:
                        print("Role is core/service",role_name)
                    else:
                        role_arn = role['Arn']
                        attached_policies = iam_client.list_attached_role_policies(RoleName=role['RoleName'])['AttachedPolicies']
                        policies_list = []
                        s3_allow_list = []
                        s3_deny_list = []
                        ec2_allow_list = []
                        ec2_deny_list = []
                        iam_allow_list = []
                        iam_deny_list = []
                        cloudtrail_allow_list = []
                        cloudtrail_deny_list = []
                        rds_allow_list = []
                        rds_deny_list = []
                        dynamodb_allow_list = []
                        dynamodb_deny_list = []
                        admin_list = []
                        compliance = " Compliance Role"
                        noncompliance = "NonCompliance Role"

                        for policy in attached_policies:
                            #policy_version = iam_client.get_policy_version(PolicyArn=policy['PolicyArn'], VersionId=policy['DefaultVersionId'])['PolicyVersion']['Document']
                            policy_document = iam_client.get_policy(PolicyArn=policy['PolicyArn'])['Policy']['DefaultVersionId']
                            #print(policy)
                            policy_version = iam_client.get_policy_version(PolicyArn=policy['PolicyArn'],VersionId=policy_document)['PolicyVersion']['Document']
                            process_policy_statement(policy_version, s3_perms_allow, s3_allow_list, s3_perms_deny, s3_deny_list, ec2_perms_allow, ec2_allow_list, ec2_perms_deny, ec2_deny_list,iam_perms_allow,iam_allow_list, iam_perms_deny, iam_deny_list, cloudtrail_perms_allow, cloudtrail_allow_list, cloudtrail_perms_deny, cloudtrail_deny_list, rds_perms_allow, rds_allow_list, rds_perms_deny, rds_deny_list, dynamodb_perms_allow, dynamodb_allow_list, dynamodb_perms_deny, dynamodb_deny_list, admin_access, admin_list)
                            policies_list.append(policy['PolicyName'])
                            if policy['PolicyName'] == 'AdministratorAccess':
                                admin_list.append("admin_access")

                        inline_policies = iam_client.list_role_policies(RoleName=role['RoleName'])['PolicyNames']
                        for policy_name in inline_policies:
                            policy_version = iam_client.get_role_policy(RoleName=role['RoleName'], PolicyName=policy_name)['PolicyDocument']
                            process_policy_statement(policy_version, s3_perms_allow, s3_allow_list, s3_perms_deny, s3_deny_list, ec2_perms_allow, ec2_allow_list, ec2_perms_deny, ec2_deny_list,iam_perms_allow,iam_allow_list, iam_perms_deny, iam_deny_list, cloudtrail_perms_allow, cloudtrail_allow_list, cloudtrail_perms_deny, cloudtrail_deny_list, rds_perms_allow, rds_allow_list, rds_perms_deny, rds_deny_list, dynamodb_perms_allow, dynamodb_allow_list, dynamodb_perms_deny, dynamodb_deny_list, admin_access, admin_list)
                            policies_list.append(policy_name)
                        total_allow = s3_allow_list + ec2_allow_list + iam_allow_list + cloudtrail_allow_list + rds_allow_list + dynamodb_allow_list + admin_list

                        if (s3_allow_list and not s3_deny_list) or (ec2_allow_list and not ec2_deny_list) or (iam_allow_list and not iam_deny_list) or (cloudtrail_allow_list and not cloudtrail_deny_list) or (rds_allow_list and not rds_deny_list) or (dynamodb_allow_list and not dynamodb_deny_list) or admin_list:
                            writer.writerow([role_name, ','.join(map(str,policies_list)), ','.join(map(str,set(s3_allow_list))), ','.join(map(str,set(s3_deny_list))), ','.join(map(str,set(ec2_allow_list))), ','.join(map(str,set(ec2_deny_list))),','.join(map(str,set(iam_allow_list))),','.join(map(str,set(iam_deny_list))),','.join(map(str,set(cloudtrail_allow_list))), ','.join(map(str,set(cloudtrail_deny_list))), ','.join(map(str,set(rds_allow_list))), ','.join(map(str,set(rds_deny_list))),','.join(map(str,set(dynamodb_allow_list))), ','.join(map(str,set(dynamodb_deny_list))),','.join(map(str,set(admin_list))), account_num,role_arn,Application, noncompliance])
                            print(role_name, f"{Application}noncompliance ---->", policies_list, s3_allow_list, s3_deny_list, ec2_allow_list, ec2_deny_list, admin_list, account_num)

                        else:
                            writer.writerow([role_name, ','.join(map(str,policies_list)), ','.join(map(str,set(s3_allow_list))), ','.join(map(str,set(s3_deny_list))), ','.join(map(str,set(ec2_allow_list))), ','.join(map(str,set(ec2_deny_list))),','.join(map(str,set(iam_allow_list))),','.join(map(str,set(iam_deny_list))),','.join(map(str,set(cloudtrail_allow_list))), ','.join(map(str,set(cloudtrail_deny_list))), ','.join(map(str,set(rds_allow_list))), ','.join(map(str,set(rds_deny_list))),','.join(map(str,set(dynamodb_allow_list))), ','.join(map(str,set(dynamodb_deny_list))),','.join(map(str,set(admin_list))), account_num,role_arn,Application, compliance])
                            print(role_name, f"{Application}compliance ---->", policies_list, s3_allow_list, s3_deny_list, ec2_allow_list, ec2_deny_list, admin_list, account_num)
    except Exception as e:
        print("Error:", e)

def process_policy_statement(policy_version, s3_perms_allow, s3_allow_list, s3_perms_deny, s3_deny_list, ec2_perms_allow, ec2_allow_list, ec2_perms_deny, ec2_deny_list,iam_perms_allow,iam_allow_list, iam_perms_deny, iam_deny_list, cloudtrail_perms_allow, cloudtrail_allow_list, cloudtrail_perms_deny, cloudtrail_deny_list, rds_perms_allow, rds_allow_list, rds_perms_deny, rds_deny_list, dynamodb_perms_allow, dynamodb_allow_list, dynamodb_perms_deny, dynamodb_deny_list, admin_access, admin_list):
    try:
      for idx, statement in enumerate(policy_version.get('Statement', []), start=1):
        effect = statement.get('Effect', 'N/A')
        actions = statement.get('Action', 'N/A')
        resources = statement.get('Resource', 'N/A')
        conditions = statement.get('Condition', 'N/A')
        #print(conditions)
        if resources.count("*") == 1 and conditions =='N/A':
            if effect == 'Allow':
                for s3_perm_allow in s3_perms_allow:
                    if s3_perm_allow in actions:
                        s3_allow_list.append(s3_perm_allow)
                for ec2_perm_allow in ec2_perms_allow:
                    if ec2_perm_allow in actions:
                        ec2_allow_list.append(ec2_perm_allow)
                for iam_perm_allow in iam_perms_allow:
                    if iam_perm_allow in actions:
                        iam_allow_list.append(iam_perm_allow)
                for cloudtrail_perm_allow in cloudtrail_perms_allow:
                    if cloudtrail_perm_allow in actions:
                        cloudtrail_allow_list.append(cloudtrail_perm_allow)
                for rds_perm_allow in rds_perms_allow:
                    if rds_perm_allow in actions:
                        rds_allow_list.append(rds_perm_allow)
                for dynamodb_perm_allow in dynamodb_perms_allow:
                    if dynamodb_perm_allow in actions:
                        dynamodb_allow_list.append(dynamodb_perm_allow)
                
            elif effect == 'Deny':
                for s3_perm_deny in s3_perms_deny:
                    if s3_perm_deny in actions:
                        s3_deny_list.append(s3_perm_deny)
                for ec2_perm_deny in ec2_perms_deny:
                    if ec2_perm_deny in actions:
                        ec2_deny_list.append(ec2_perm_deny)
                for iam_perm_deny in iam_perms_deny:
                    if iam_perm_deny in actions:
                        iam_deny_list.append(iam_perm_deny)
                for cloudtrail_perm_deny in cloudtrail_perms_deny:
                    if cloudtrail_perm_deny in actions:
                        cloudtrail_deny_list.append(cloudtrail_perm_deny)
                for rds_perm_deny in rds_perms_deny:
                    if rds_perm_deny in actions:
                        rds_deny_list.append(rds_perm_deny)
                for dynamodb_perm_deny in dynamodb_perms_deny:
                    if dynamodb_perm_deny in actions:
                        dynamodb_deny_list.append(dynamodb_perm_deny)
            
    except Exception as e1:
      print(e1)
      pass
if __name__ == "__main__":
    main()

roles.py
from multiprocessing.sharedctypes import Value
import boto3
from datetime import date
import xlwings as xw
import openpyxl
import botocore
import pandas as pd
import json
from botocore.exceptions import ClientError
def main():
    sts_client = boto3.client('sts')
    #file_name= 'Accesskeys_'+str(date.today())+".csv"
    with open("roles_report.csv",'w') as f: 
        f.write("RoleName"+","+"RoleId"+","+"Arn"+","+"CreateDate"+","+"LastUsedDate"+","+"PolicyNames"+","+"Trust Relationship"+","+"PolicyDocument"+","+"AccountNumber\n")
        #f.write("RoleName"+","+"RoleId"+","+"Arn"+","+"CreateDate"+","+"LastUsedDate"+","+"Role Age"+","+"AccountNumber\n")
        assumed_role_object = sts_client.assume_role(RoleArn='arn:aws:iam::528150397796:role/IAM_Automation',RoleSessionName="iamautomation",ExternalId="UZZCWIWGSY")
        credentials=assumed_role_object['Credentials']
        client=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
        root_acc = client.get_caller_identity().get('Account')
        print(root_acc)
        assumerole_iam(client,f,sts_client)
def assumerole_iam(client,f,sts_client):
        assumed_role_object = sts_client.assume_role(RoleArn="arn:aws:iam::528150397796:role/528150397796-Devops-Admin-Role",RoleSessionName="iamautomation")
        credentials=assumed_role_object['Credentials']
        iam_client=boto3.client('iam',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],
        aws_session_token=credentials['SessionToken'])
        sts_clientroot1=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
        root_acc = sts_clientroot1.get_caller_identity()
        print(root_acc)
        fname = 'roles-dev.xlsx'
        wb = openpyxl.load_workbook(fname)
        sheet = wb.get_sheet_by_name('Sheet1')
        for rowOfCellObjects in sheet['A71':'A110']:
            try:
                for cellObj in rowOfCellObjects:
                    #print(cellObj.coordinate, cellObj.value)
                    v1 = cellObj.value
                    assumed_role_object = sts_clientroot1.assume_role(RoleArn=v1,RoleSessionName="AssumeRoleSession1")
                    credentials=assumed_role_object['Credentials']
                    iam_client=boto3.client('iam',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],
                    aws_session_token=credentials['SessionToken'])
                    sts_client1=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],
                    aws_session_token=credentials['SessionToken'])
                    account_num = sts_client1.get_caller_identity().get('Account')
                    #acc_name =   boto3.client('organizations').describe_account(AccountId=account_num).get('Account').get('Name')
                    users = iam_client.list_users()["Users"]
                    print("Switching to",account_num)
                    iam_roles(iam_client,client,f,account_num)
            except botocore.exceptions.ClientError as error: 
                print(error)
                continue
def iam_roles(iam_client,client,f,account_num):    
    paginator = iam_client.get_paginator("list_roles")
    for page in paginator.paginate(PaginationConfig={'MaxItems': 1000}):
        for listed_role in page["Roles"]:
            role_name = listed_role["RoleName"]
            role_id = listed_role["RoleId"]
            role_arn = listed_role["Arn"]
            role = iam_client.get_role(RoleName=role_name)["Role"]
            last_used = role.get("RoleLastUsed", {}).get("LastUsedDate")
            role_doc= role.get('AssumeRolePolicyDocument')
            assume_document = str(role_doc).replace(",","$")
            lastuseddate = str(last_used)
            created = listed_role["CreateDate"].date()
            currentdate = date.today()
            roleage = currentdate - created
            #print(roleage.days)
            createddate = str(created)
            policy = iam_client.list_role_policies(RoleName=role_name)['PolicyNames']
            for policy_name in policy:
              role_policy = iam_client.get_role_policy(RoleName=role_name,PolicyName=policy_name)
              print(role_policy)
              policy_document = role_policy['PolicyDocument']
              policy_docum = str(policy_document).replace(",","$")
              f.write(role_name+","+role_id+","+role_arn+","+createddate+","+lastuseddate+","+policy_name+","+assume_document+","+policy_docum+","+account_num+"\n")
            policy_name = str(policy)
            #print(policy)
            policy_names = policy_name.replace(",","$")
            cus_managed = iam_client.list_attached_role_policies(RoleName=role_name)['AttachedPolicies']
            length = len(cus_managed)
            try:
                for i in range(length):   
                    polarn = cus_managed[i]["PolicyArn"]
                    polname = cus_managed[i]['PolicyName']
                    response = iam_client.get_policy(PolicyArn=polarn)
                    policy_version = iam_client.get_policy_version( PolicyArn = polarn, VersionId = response['Policy']['DefaultVersionId'] )
                    
					#Hashed earlier f.write(json.dumps(policy_version['PolicyVersion']['Document']))
                    docum = policy_version['PolicyVersion']['Document']
                    document = str(docum).replace(",","$")
                    
                    f.write(role_name+","+role_id+","+role_arn+","+createddate+","+lastuseddate+","+polname+","+assume_document+","+document+","+account_num+"\n")
                    #f.write(role_name+","+role_id+","+role_arn+","+createddate+","+lastuseddate+","+str(roleage.days)+","+account_num+"\n")
                    
                    #print(json.dumps(policy_version['PolicyVersion']['Document']['Statement']))
            except Exception as e:
                print(e)
                continue
if __name__ == "__main__":
    main()

roles_deletion.py
from multiprocessing.sharedctypes import Value
import boto3
from datetime import date
import xlwings as xw
import openpyxl
import botocore
import pandas as pd
import json
from botocore.exceptions import ClientError
def main():
    sts_client = boto3.client('sts')
    assumed_role_object = sts_client.assume_role(RoleArn="arn:aws:iam::528150397796:role/528150397796-Devops-Admin-Role",RoleSessionName="iamautomation")
    credentials=assumed_role_object['Credentials']
    devops_sts = boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
    iam_client=boto3.client('iam',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
    root_acc = sts_client.get_caller_identity().get('Account')
    print(root_acc)
    list_077249158321 = ['077249158321-SRE-Admin-Role']
    list_142222938030 = ['142222938030-SRE-Admin-Role']
    list_156744138451 = ['156744138451-SRE-Admin-Role']
    list_242856619680 = ['242856619680-SRE-Admin-Role']
    list_306267571650 = ['306267571650-SRE-Admin-Role']
    list_367742456716 = ['367742456716-SRE-Admin-Role']
    list_422880992407 = ['422880992407-SRE-Admin-Role']
    list_435953702003 = ['435953702003-SRE-Admin-Role']
    list_463942034412 = ['463942034412-SRE-Admin-Role']
    list_504397588686 = ['504397588686-SRE-Admin-Role']
    list_506034490444 = ['506034490444-SRE-Admin-Role']
    list_554554622031 = ['554554622031-SRE-Admin-Role']
    list_574957753581 = ['574957753581-SRE-Admin-Role']
    list_612164320879 = ['612164320879-SRE-Admin-Role']
    list_633520970819 = ['633520970819-SRE-Admin-Role']
    list_662012991696 = ['662012991696-SRE-Admin-Role']
    list_729303886551 = ['729303886551-SRE-Admin-Role']
    list_832601267655 = ['832601267655-SRE-Admin-Role']
    list_891853622509 = ['891853622509-SRE-Admin-Role']
    list_963696105088 = ['963696105088-SRE-Admin-Role']
    list_992382813542 = ['992382813542-SRE-Admin-Role']
    

    lists = {
        '077249158321':	list_077249158321,
        '142222938030': list_142222938030,
        '156744138451':	list_156744138451,
        '242856619680':	list_242856619680,
        '306267571650':	list_306267571650,
        '367742456716':	list_367742456716,
        '422880992407':	list_422880992407,
        '435953702003':	list_435953702003,
        '463942034412':	list_463942034412,
        '504397588686':	list_504397588686,
        '506034490444':	list_506034490444,
        '554554622031':	list_554554622031,
        '574957753581':	list_574957753581,
        '612164320879':	list_612164320879,
        '633520970819':	list_633520970819,
        '662012991696':	list_662012991696,
        '729303886551':	list_729303886551,
        '832601267655':	list_832601267655,
        '891853622509':	list_891853622509,
        '963696105088':	list_963696105088,
        '992382813542':	list_992382813542
    }
    rolestoswitch = [
      "arn:aws:iam::077249158321:role/077249158321-Devops-Admin-Role",
      "arn:aws:iam::142222938030:role/142222938030-Devops-Admin-Role",
      "arn:aws:iam::156744138451:role/156744138451-Devops-Admin-Role",
      "arn:aws:iam::242856619680:role/242856619680-Devops-Admin-Role",
      "arn:aws:iam::306267571650:role/306267571650-Devops-Admin-Role",
      "arn:aws:iam::367742456716:role/367742456716-Devops-Admin-Role",
      "arn:aws:iam::422880992407:role/422880992407-Devops-Admin-Role",
      "arn:aws:iam::435953702003:role/435953702003-Devops-Admin-Role",
      "arn:aws:iam::463942034412:role/463942034412-Devops-Admin-Role",
      "arn:aws:iam::504397588686:role/504397588686-Devops-Admin-Role",
      "arn:aws:iam::506034490444:role/506034490444-Devops-Admin-Role",
      "arn:aws:iam::554554622031:role/554554622031-Devops-Admin-Role",
      "arn:aws:iam::574957753581:role/574957753581-Devops-Admin-Role",
      "arn:aws:iam::612164320879:role/612164320879-Devops-Admin-Role",
      "arn:aws:iam::633520970819:role/633520970819-Devops-Admin-Role",
      "arn:aws:iam::662012991696:role/662012991696-Devops-Admin-Role",
      "arn:aws:iam::729303886551:role/729303886551-Devops-Admin-Role",
      "arn:aws:iam::832601267655:role/832601267655-Devops-Admin-Role",
      "arn:aws:iam::891853622509:role/891853622509-Devops-Admin-Role",
      "arn:aws:iam::963696105088:role/963696105088-Devops-Admin-Role",
      "arn:aws:iam::992382813542:role/992382813542-Devops-Admin-Role"
	]
    for role in rolestoswitch:
        try:
            assumed_role_object = devops_sts.assume_role(RoleArn=role,RoleSessionName="AssumeRoleSession1")
            credentials=assumed_role_object['Credentials']
            iam=boto3.client('iam',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
            sts_clientroot1=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
            account_num = sts_clientroot1.get_caller_identity().get('Account')
            print(account_num)
            print("Switching to",account_num)
            foundlist = lists.get(account_num)
            for role_name in foundlist:
                try:
                    response = iam.get_role(RoleName=role_name)
                    role_arn = response['Role']['Arn']
                        
                    attached_policies = iam.list_attached_role_policies(RoleName=role_name)['AttachedPolicies']
                    for policy in attached_policies:
                        iam.detach_role_policy(RoleName=role_name, PolicyArn=policy['PolicyArn'])
                        
                    inline_policies = iam.list_role_policies(RoleName=role_name)['PolicyNames']
                    for policy_name in inline_policies:
                        iam.delete_role_policy(RoleName=role_name, PolicyName=policy_name)
                        
                    iam.delete_role(RoleName=role_name)
                        
                    print(f"IAM role {role_arn} has been removed.")
                except Exception as e:
                    print(e)
                    pass
        except botocore.exceptions.ClientError as error: 
                print(error)
                continue

if __name__ == "__main__":
	main()

roles_lastused.py
from multiprocessing.sharedctypes import Value
import boto3
from datetime import date
import xlwings as xw
import openpyxl
import botocore
import pandas as pd
import json
from botocore.exceptions import ClientError
def main():
    sts_client = boto3.client('sts')
    #file_name= 'Accesskeys_'+str(date.today())+".csv"
    rlu = open("RoleLastUsed.csv","w")
    servroles = open("ServiceRoles.csv","w")
    servroles.write("RoleName"+","+"RoleId"+","+"Arn"+","+"CreateDate"+","+"LastUsedDate"+","+"PolicyNames"+","+"Trust Relationship"+","+"PolicyDocument"+","+"AccountNumber\n")
    rlu.write("RoleName"+","+"RoleId"+","+"Arn"+","+"CreateDate"+","+"LastUsedDate"+","+"PolicyNames"+","+"Trust Relationship"+","+"PolicyDocument"+","+"AccountNumber\n")
    with open("roles_report.csv",'w') as f: 
        f.write("RoleName"+","+"RoleId"+","+"Arn"+","+"CreateDate"+","+"LastUsedDate"+","+"PolicyNames"+","+"Trust Relationship"+","+"PolicyDocument"+","+"AccountNumber\n")
        assumed_role_object = sts_client.assume_role(RoleArn='arn:aws:iam::528150397796:role/IAM_Automation',RoleSessionName="iamautomation",ExternalId="UZZCWIWGSY")
        credentials=assumed_role_object['Credentials']
        client=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
        root_acc = client.get_caller_identity().get('Account')
        print(root_acc)
        assumerole_iam(client,f,rlu,servroles)
    rlu.close()
    servroles.close()
def assumerole_iam(client,f,rlu,servroles):
        assumed_role_object = client.assume_role(RoleArn="arn:aws:iam::528150397796:role/528150397796-Devops-Admin-Role",RoleSessionName="iamautomation")
        credentials=assumed_role_object['Credentials']
        iam_client=boto3.client('iam',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],
        aws_session_token=credentials['SessionToken'])
        sts_clientroot1=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
        root_acc = sts_clientroot1.get_caller_identity()
        print(root_acc)
        fname = 'roles-dev.xlsx'
        wb = openpyxl.load_workbook(fname)
        sheet = wb.get_sheet_by_name('Sheet1')
        for rowOfCellObjects in sheet['A1':'A96']:
            try:
                for cellObj in rowOfCellObjects:
                    #print(cellObj.coordinate, cellObj.value)
                    v1 = cellObj.value
                    assumed_role_object = sts_clientroot1.assume_role(RoleArn=v1,RoleSessionName="AssumeRoleSession1")
                    credentials=assumed_role_object['Credentials']
                    iam_client=boto3.client('iam',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],
                    aws_session_token=credentials['SessionToken'])
                    sts_client1=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],
                    aws_session_token=credentials['SessionToken'])
                    account_num = sts_client1.get_caller_identity().get('Account')
                    #acc_name =   boto3.client('organizations').describe_account(AccountId=account_num).get('Account').get('Name')
                    users = iam_client.list_users()["Users"]
                    print("Switching to",account_num)
                    iam_roles(iam_client,client,f,account_num,rlu,servroles)
            except botocore.exceptions.ClientError as error: 
                print(error)
                continue
def iam_roles(iam_client,client,f,account_num,rlu,servroles):    
    paginator = iam_client.get_paginator("list_roles")
    for page in paginator.paginate():
        for listed_role in page["Roles"]:
            role_name = listed_role["RoleName"]
            role_id = listed_role["RoleId"]
            role_arn = listed_role["Arn"]
            role = iam_client.get_role(RoleName=role_name)["Role"]
            last_used = role.get("RoleLastUsed", {}).get("LastUsedDate")
            role_doc= role.get('AssumeRolePolicyDocument')
            assume_document = str(role_doc).replace(",","$")
            lastuseddate = str(last_used)
            created = listed_role["CreateDate"].date()
            createddate = str(created)
            policy = iam_client.list_role_policies(RoleName=role_name)['PolicyNames']
            policy_name = str(policy)
            #print(policy)
            policy_names = policy_name.replace(",","$")
            cus_managed = iam_client.list_attached_role_policies(RoleName=role_name)['AttachedPolicies']
            length = len(cus_managed)
            try:
                for i in range(length):   
                    polarn = cus_managed[i]["PolicyArn"]
                    polname = cus_managed[i]['PolicyName']
                    response = iam_client.get_policy(PolicyArn=polarn)
                    policy_version = iam_client.get_policy_version( PolicyArn = polarn, VersionId = response['Policy']['DefaultVersionId'] )
                    #f.write(json.dumps(policy_version['PolicyVersion']['Document']))
                    docum = policy_version['PolicyVersion']['Document']
                    document = str(docum).replace(",","$")
                    #f.write(role_name+","+role_id+","+role_arn+","+createddate+","+lastuseddate+","+polname+","+assume_document+","+document+","+account_num+"\n")
                    tagvalues = role.get("Tags")
                    try:
                        for tagvalue in tagvalues:
                          #if tagvalue['Key']=='Function' and tagvalue['Value']=="Service-Role":
                            #servroles.write(role_name+","+role_id+","+role_arn+","+createddate+","+lastuseddate+","+polname+","+assume_document+","+document+","+account_num+"\n")
                          if tagvalue['Key']=='Name':
                            servroles.write(role_name+","+role_id+","+role_arn+","+account_num+","+"\n")
                            print(tagvalue['Key'],tagvalue['Value'])
                    except TypeError:
                      print('TypeError')
                      #servroles.write(role_name+","+role_id+","+role_arn+","+account_num+","+"Notags\n")
                      continue
                    #print(json.dumps(policy_version['PolicyVersion']['Document']['Statement']))
                    try:
                        last_used_test = role.get("RoleLastUsed")
                        #print(last_used_test["LastUsedDate"].date())
                        currentdate = date.today()
                        last_used_sub = currentdate - last_used_test["LastUsedDate"].date()
                        #print("Last Used Date: \t",last_used_sub.days)
                        rlu.write(role_name+","+role_id+","+role_arn+","+createddate+","+str(last_used_sub.days)+","+polname+","+assume_document+","+document+","+account_num+"\n")
                    except KeyError:
                        #print("NoRLU")
                        rlu.write(role_name+","+role_id+","+role_arn+","+createddate+","+"nil"+","+polname+","+assume_document+","+document+","+account_num+"\n")
                        continue
            except Exception as e:
                print(e)
                continue
if __name__ == "__main__":
    main()


roles_tagging.py
from multiprocessing.sharedctypes import Value
import boto3
from datetime import date
import xlwings as xw
import openpyxl
import botocore
import pandas as pd
import json
from botocore.exceptions import ClientError
def main():
  sts_client = boto3.client('sts')
  assumed_role_object = sts_client.assume_role(RoleArn='arn:aws:iam::528150397796:role/IAM_Automation',RoleSessionName="iamautomation",ExternalId="UZZCWIWGSY")
  credentials=assumed_role_object['Credentials']
  client=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
  root_acc = client.get_caller_identity().get('Account')
  print(root_acc)
  assumerole_iam(client)
def assumerole_iam(client):
  assumed_role_object = client.assume_role(RoleArn="arn:aws:iam::528150397796:role/528150397796-Devops-Admin-Role",RoleSessionName="iamautomation")
  credentials=assumed_role_object['Credentials']
  iam_client=boto3.client('iam',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
  sts_clientroot1=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
  root_acc = sts_clientroot1.get_caller_identity()
  print(root_acc)
  fname = 'Dev-accounts.xlsx'
  wb = openpyxl.load_workbook(fname)
  sheet = wb.get_sheet_by_name('Sheet1')
  for rowOfCellObjects in sheet['A2':'A42']:
    try:
      for cellObj in rowOfCellObjects:
        #print(cellObj.coordinate, cellObj.value)
        v1 = cellObj.value
        assumed_role_object = sts_clientroot1.assume_role(RoleArn=v1,RoleSessionName="AssumeRoleSession1")
        credentials=assumed_role_object['Credentials']
        iam_client=boto3.client('iam',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
        sts_client1=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
        account_num = sts_client1.get_caller_identity().get('Account')
        #acc_name =   boto3.client('organizations').describe_account(AccountId=account_num).get('Account').get('Name') 
        additional_tags = [{'Key': 'role_function', 'Value': 'service'}]
        additional_tags1 = [{'Key': 'role_function', 'Value': 'core'}]
        f = open("issues-service.csv",'w')
        f1 = open("issues-core.csv",'w')
        print("Switching to",account_num)
        service_roles(iam_client,additional_tags,account_num,f)
        #core_roles(iam_client,additional_tags1,account_num,f1)
    except botocore.exceptions.ClientError as error:
      print(error)
      continue  
def service_roles(iam_client,additional_tags,account_num,f):
  role_names =['OrganizationAccountAccessRole','expn-cis-config-role']
  #role_names =['rds-monitoring-role','CloudabilityRole','WizAccess-Role','Okta-Idp-cross-account-role','stacksets-exec-c1c1b0535f75d712e3199a0026442703']
  #role_names =['Okta-Idp-cross-account-role','stacksets-exec-c1c1b0535f75d712e3199a0026442703']
  #role_names =['WizAccess-Role']
  #role_names =['stacksets-exec-c1c1b0535f75d712e3199a0026442703']
  #cmr_role=account_num+"-cloud-management-Role"
  #role_names.append(cmr_role)
  try:
    for role_name in role_names:
      response = iam_client.list_role_tags(RoleName=role_name)
      current_tags = response['Tags']
      # Add additional tags to the existing ones
      updated_tags = current_tags + additional_tags
      # Update tags for the IAM role
      iam_client.tag_role(RoleName=role_name, Tags=updated_tags)
      print(f"Additional tags added to IAM role {role_name} successfully.")
  except botocore.exceptions.ClientError as error:
    print(error)
    f.write(role_name+","+account_num+","+str(error)+"\n")
    pass
        
def core_roles(iam_client,additional_tags1,account_num,f1):
  try:
    role_name = account_num+"-DBE-Role"
    response = iam_client.list_role_tags(RoleName=role_name)
    current_tags = response['Tags']
    # Add additional tags to the existing ones
    updated_tags = current_tags + additional_tags1
    # Update tags for the IAM role
    iam_client.tag_role(RoleName=role_name, Tags=updated_tags)
    print(f"Additional tags added to IAM role {role_name} successfully.")
  except botocore.exceptions.ClientError as error:
    print(error)
    f1.write(role_name+","+account_num+","+str(error)+"\n")
    pass
if __name__ == "__main__":
    main()


rolestags.py
from multiprocessing.sharedctypes import Value
import boto3
from datetime import date
import xlwings as xw
import openpyxl
import botocore
import pandas as pd
import json
from botocore.exceptions import ClientError
def main():
    sts_client = boto3.client('sts')
    #file_name= 'Accesskeys_'+str(date.today())+".csv"
    with open("roles_report.csv",'w') as f: 
        f.write("RoleName"+","+"RoleId"+","+"Arn"+","+"CreateDate"+","+"LastUsedDate"+","+"PolicyNames"+","+"Trust Relationship"+","+"PolicyDocument"+","+"AccountNumber\n")
        #f.write("RoleName"+","+"RoleId"+","+"Arn"+","+"CreateDate"+","+"LastUsedDate"+","+"Role Age"+","+"AccountNumber\n")
        assumed_role_object = sts_client.assume_role(RoleArn='arn:aws:iam::528150397796:role/IAM_Automation',RoleSessionName="iamautomation",ExternalId="UZZCWIWGSY")
        credentials=assumed_role_object['Credentials']
        client=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
        root_acc = client.get_caller_identity().get('Account')
        print(root_acc)
        assumerole_iam(client,f)
def assumerole_iam(client,f):
        assumed_role_object = client.assume_role(RoleArn="arn:aws:iam::528150397796:role/528150397796-Devops-Admin-Role",RoleSessionName="iamautomation")
        credentials=assumed_role_object['Credentials']
        iam_client=boto3.client('iam',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],
        aws_session_token=credentials['SessionToken'])
        sts_clientroot1=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
        root_acc = sts_clientroot1.get_caller_identity()
        print(root_acc)
        fname = 'roles-dev.xlsx'
        wb = openpyxl.load_workbook(fname)
        sheet = wb.get_sheet_by_name('Sheet1')
        for rowOfCellObjects in sheet['A1':'A103']:
            try:
                for cellObj in rowOfCellObjects:
                    #print(cellObj.coordinate, cellObj.value)
                    v1 = cellObj.value
                    assumed_role_object = sts_clientroot1.assume_role(RoleArn=v1,RoleSessionName="AssumeRoleSession1")
                    credentials=assumed_role_object['Credentials']
                    iam_client=boto3.client('iam',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],
                    aws_session_token=credentials['SessionToken'])
                    sts_client1=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],
                    aws_session_token=credentials['SessionToken'])
                    account_num = sts_client1.get_caller_identity().get('Account')
                    #acc_name =   boto3.client('organizations').describe_account(AccountId=account_num).get('Account').get('Name')
                    users = iam_client.list_users()["Users"]
                    print("Switching to",account_num)
                    iam_roles(iam_client,client,f,account_num)
            except botocore.exceptions.ClientError as error: 
                print(error)
                continue
def iam_roles(iam_client,client,f,account_num):    
    paginator = iam_client.get_paginator("list_roles")
    for page in paginator.paginate(PaginationConfig={'MaxItems': 1000}):
        for listed_role in page["Roles"]:
            role_name = listed_role["RoleName"]
            role_id = listed_role["RoleId"]
            arn = listed_role["Arn"]
            create_date = listed_role["CreateDate"]
            tags = iam_client.list_role_tags(RoleName=role_name)
            for tag in tags['Tags']:
              print(tag['Key'],tag['Value'])
              f.write(role_name+","+arn+","+tag['Key']+","+tag['Value']+","+account_num+","+"\n")
            
if __name__ == "__main__":
    main()

rolesvalidation.py
import boto3
from datetime import date
import csv
import openpyxl
from botocore.exceptions import ClientError

def main():
    # Assume the IAM Automation Role to start
    sts_client = boto3.client('sts')
    assumed_role_object = sts_client.assume_role(
        RoleArn='arn:aws:iam::528150397796:role/528150397796-Devops-Admin-Role',
        RoleSessionName="iamautomation"
    )
    credentials = assumed_role_object['Credentials']
    client = boto3.client(
        'sts',
        aws_access_key_id=credentials['AccessKeyId'],
        aws_secret_access_key=credentials['SecretAccessKey'],
        aws_session_token=credentials['SessionToken']
    )
    
    # Get the root account number
    root_acc = client.get_caller_identity().get('Account')
    print("Root Account:", root_acc)
    
    # Open CSV file for writing
    with open("roles_report.csv", 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["RoleName", "RoleArn", "LastUsedDate", "AccountNumber"])
        
        # Call the function to assume and gather roles information from Excel
        assumerole_iam(client, writer, sts_client)

def assumerole_iam(client, writer, sts_client):
    # Load the Excel file and sheet
    fname = 'roles.xlsx'
    wb = openpyxl.load_workbook(fname)
    sheet = wb['Sheet1']
    
    # Loop through the specified range of cells for role ARNs
    for row in sheet['A1':'A110']:
        for cell in row:
            role_arn = cell.value
            if role_arn:
                try:
                    # Assume role for each ARN found in the Excel sheet
                    assumed_role_object = client.assume_role(
                        RoleArn=role_arn,
                        RoleSessionName="AssumeRoleSession1"
                    )
                    credentials = assumed_role_object['Credentials']
                    
                    iam_client = boto3.client(
                        'iam',
                        aws_access_key_id=credentials['AccessKeyId'],
                        aws_secret_access_key=credentials['SecretAccessKey'],
                        aws_session_token=credentials['SessionToken']
                    )
                    
                    sts_client1 = boto3.client(
                        'sts',
                        aws_access_key_id=credentials['AccessKeyId'],
                        aws_secret_access_key=credentials['SecretAccessKey'],
                        aws_session_token=credentials['SessionToken']
                    )
                    
                    account_num = sts_client1.get_caller_identity().get('Account')
                    print(f"Switching to Account: {account_num}")
                    
                    # Collect role data and write to CSV
                    iam_roles(iam_client, writer, account_num)
                
                except ClientError as error:
                    print(f"Error assuming role for {role_arn}: {error}")
                    continue

def iam_roles(iam_client, writer, account_num):
    paginator = iam_client.get_paginator("list_roles")
    
    for page in paginator.paginate(PaginationConfig={'MaxItems': 1000}):
        for listed_role in page["Roles"]:
            role_name = listed_role["RoleName"]
            role_arn = listed_role["Arn"]
            role_used_get = iam_client.get_role(RoleName=role_name)["Role"]
            last_used = role_used_get.get("RoleLastUsed", {}).get("LastUsedDate")

            # Write the required information to CSV
            writer.writerow([role_name, role_arn, str(last_used), account_num])
            print(role_name,role_arn,str(last_used), account_num)

if __name__ == "__main__":
    main()


S3kms.py

from multiprocessing.sharedctypes import Value
import boto3
from datetime import date
import xlwings as xw
import openpyxl
import botocore
import pandas as pd
import json
from botocore.exceptions import ClientError
def main():
    sts_client = boto3.client('sts')
    with open("s3_enc"+'.csv','w') as f: 
        f.write("S3 Bucket Name"+","+"Encryption Algorithm"+","+"BucketKeyEnabled"+","+"KMSMasterKeyID"+","+"AccountNum\n")
        #assumed_role_object = sts_client.assume_role(RoleArn='arn:aws:iam::528150397796:role/IAM_Automation',RoleSessionName="iamautomation",ExternalId="UZZCWIWGSY")
        #credentials=assumed_role_object['Credentials']
        #client=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
        #root_acc = client.get_caller_identity().get('Account')
        #print(root_acc)
        assumerole_iam(sts_client,f)
def assumerole_iam(sts_client,f):
        assumed_role_object = sts_client.assume_role(RoleArn="arn:aws:iam::528150397796:role/528150397796-Devops-Admin-Role",RoleSessionName="iamautomation")
        credentials=assumed_role_object['Credentials']
        iam_client=boto3.client('iam',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],
        aws_session_token=credentials['SessionToken'])
        sts_clientroot1=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],aws_session_token=credentials['SessionToken'])
        root_acc = sts_clientroot1.get_caller_identity()
        print(root_acc)
        fname = 'roles-dev.xlsx'
        wb = openpyxl.load_workbook(fname)
        sheet = wb.get_sheet_by_name('Sheet1')
        for rowOfCellObjects in sheet['A1':'A103']:
            try:
                for cellObj in rowOfCellObjects:
                    #print(cellObj.coordinate, cellObj.value)
                    v1 = cellObj.value
                    assumed_role_object = sts_clientroot1.assume_role(RoleArn=v1,RoleSessionName="AssumeRoleSession1")
                    credentials=assumed_role_object['Credentials']
                    s3_client=boto3.client('s3',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],
                    aws_session_token=credentials['SessionToken'])
                    sts_client1=boto3.client('sts',aws_access_key_id=credentials['AccessKeyId'],aws_secret_access_key=credentials['SecretAccessKey'],
                    aws_session_token=credentials['SessionToken'])
                    account_num = sts_client1.get_caller_identity().get('Account')
                    #acc_name =   boto3.client('organizations').describe_account(AccountId=account_num).get('Account').get('Name')
                    #users = iam_client.list_users()["Users"]
                    print("Switching to",account_num)
                    s3buckets(s3_client,f,account_num)
            except botocore.exceptions.ClientError as error: 
                print(error)
                continue
def s3buckets(s3_client,f,account_num):    
        response = s3_client.list_buckets()
        for s3name in response["Buckets"]:
            try:
                response_enc = s3_client.get_bucket_encryption(Bucket=s3name['Name'])
                #response = s3_client.get_bucket_tagging(Bucket='string')
                for s3_status in response_enc['ServerSideEncryptionConfiguration']['Rules']:
                    s3_algorithm = s3_status['ApplyServerSideEncryptionByDefault']['SSEAlgorithm']
                    s3_KeyEnabled = s3_status['BucketKeyEnabled']
                    #print(s3name["Name"],s3_status['ApplyServerSideEncryptionByDefault']['SSEAlgorithm'],s3_status['BucketKeyEnabled'])
                    try:
                        print(s3name["Name"],s3_status['ApplyServerSideEncryptionByDefault']['KMSMasterKeyID'])
                        f.write(s3name['Name']+","+s3_algorithm+","+str(s3_KeyEnabled)+","+s3_status['ApplyServerSideEncryptionByDefault']['KMSMasterKeyID']+","+account_num+"\n")
                        #response = kms_client.describe_key(KeyId = s3_status['ApplyServerSideEncryptionByDefault']['KMSMasterKeyID'])
                        print(response)
                    except KeyError:
                        print("No KMS Key Available")
                        f.write(s3name['Name']+","+s3_algorithm+","+str(s3_KeyEnabled)+","+" "+","+account_num+"\n")
                        continue
            except ClientError as e:
                if e.response['Error']['Code'] == 'ServerSideEncryptionConfigurationNotFoundError':
                    print('Bucket: %s, no server-side encryption' % (s3name['Name']))
                    f.write(s3name['Name']+","+"NoEncryptionalgorithm"+","+"Nos3KeyEnabled"+","+"NoKMSKeyID"+","+account_num+"\n")
                    continue
if __name__ == "__main__":
    main()

tfroles.py
import os
import openpyxl
from subprocess import PIPE, run
from pandas import *
import re
with open("tfroles1.csv",'w') as f:
    roleid = read_csv("rolespart1.csv")
    roleids = roleid['RoleId'].tolist()
    accnum = roleid['accnum'].tolist()
    i =0
    while i < len(roleids):
        #status = os.system(f"egrep -i {roleids[i]} {accnum[i]}/*.tfstate")
        cmd = f'awk \'/"unique_id": "{roleids[i]}"/\' {accnum[i]}/*.tfstate'
        result = run(cmd, stdout=PIPE, stderr=PIPE, universal_newlines=True, shell=True)
        if len(result.stdout) == 0:
            print(result.stdout)
        else:
            print("The output is :", result.stdout)
            f.write(roleids[i]+","+str(accnum[i])+","+result.stdout)
        #os.system(output = $(cmd))
        i = i+1

